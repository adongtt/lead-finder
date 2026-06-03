#!/usr/bin/env python3
"""
展会参展商提取 + 手套客户筛选工具

用途：从展会官网提取参展商名单，自动筛选手套相关公司，并查找官网域名。
输出可直接复制到 lead-finder 的"批量导入域名"功能中使用。

支持三种输入方式：
  1. 网页 URL: 自动抓取页面上的公司名列表
  2. Excel/CSV 文件: 读取已有参展商名单
  3. 文本列表: 直接粘贴公司名（每行一个）

用法示例：
  # 从网页抓取
  python extract_trade_show_leads.py --url "https://ispo.com/en/exhibitors" --output exhibitors.csv

  # 从 Excel 读取
  python extract_trade_show_leads.py --file "nsga_exhibitors.xlsx" --col "Company Name" --output leads.csv

  # 从文本读取
  python extract_trade_show_leads.py --text-file "companies.txt" --output leads.csv

  # 查找域名（对已有公司名列表）
  python extract_trade_show_leads.py --file "exhibitors.xlsx" --find-domains --output leads.csv
"""

import argparse
import csv
import re
import sys
import time
import urllib.parse
from pathlib import Path
from typing import List, Optional, Set

try:
    import requests
except ImportError:
    print("[ERROR] requests not installed. Run: pip install requests")
    sys.exit(1)

try:
    from bs4 import BeautifulSoup
except ImportError:
    BeautifulSoup = None

try:
    import openpyxl
except ImportError:
    openpyxl = None

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    )
}

# Keywords that indicate a glove-related company
GLOVE_KEYWORDS = [
    "glove", "gloves",
    "hand protection", "handwear",
    "baseball", "football", "softball", "lacrosse", "hockey",
    "soccer", "goalkeeper", "goalie",
    "ski", "snowboard", "winter sport",
    "cycling", "bike", "motorcycle",
    "work", "safety", "industrial", "construction",
    "gym", "fitness", "weightlifting",
    "boxing", "mma", "martial art",
    "batting", "fielding",
    "goal keeping", "goalkeeping",
]

# ---------------------------------------------------------------------------
# Text extraction helpers
# ---------------------------------------------------------------------------

def normalize_company_name(name: str) -> str:
    """Clean up a company name for consistent matching."""
    name = name.strip()
    # Remove common suffixes
    name = re.sub(r"\s+(Inc\.?|LLC|Ltd\.?|Limited|Corp\.?|Corporation|GmbH|AG|S\.A\.?|B\.V\.?)$", "", name, flags=re.IGNORECASE)
    return name.strip()


def is_glove_related(text: str) -> bool:
    """Check if text contains glove-related keywords."""
    text_lower = text.lower()
    return any(kw in text_lower for kw in GLOVE_KEYWORDS)


# ---------------------------------------------------------------------------
# Web scraping
# ---------------------------------------------------------------------------

def fetch_html(url: str, timeout: int = 30) -> str:
    """Fetch page HTML."""
    try:
        resp = requests.get(url, headers=HEADERS, timeout=timeout)
        resp.raise_for_status()
        return resp.text
    except Exception as e:
        print(f"[ERROR] Failed to fetch {url}: {e}")
        return ""


def extract_company_names_from_html(html: str) -> List[str]:
    """
    Extract potential company names from HTML.
    Uses multiple heuristics to find exhibitor/company lists.
    """
    if not html:
        return []

    names: Set[str] = set()

    if BeautifulSoup:
        soup = BeautifulSoup(html, "html.parser")

        # Heuristic 1: Look for common exhibitor list structures
        # Pattern: div/li with class containing "exhibitor", "company", "participant"
        for cls in ["exhibitor", "company", "participant", "vendor", "supplier", "delegate"]:
            for elem in soup.find_all(attrs={"class": re.compile(cls, re.I)}):
                text = elem.get_text(strip=True)
                if text and len(text) < 100 and len(text) > 2:
                    names.add(text)

        # Heuristic 2: h2/h3/h4 tags inside list items or cards
        for tag in ["h2", "h3", "h4", "strong"]:
            for elem in soup.find_all(tag):
                text = elem.get_text(strip=True)
                if text and 3 < len(text) < 80:
                    names.add(text)

        # Heuristic 3: table rows with company names
        for row in soup.find_all("tr"):
            cells = row.find_all(["td", "th"])
            if cells:
                text = cells[0].get_text(strip=True)
                if text and 3 < len(text) < 80:
                    names.add(text)

        # Heuristic 4: aria-label or data-name attributes
        for elem in soup.find_all(attrs={"data-name": True}):
            names.add(elem["data-name"].strip())
        for elem in soup.find_all(attrs={"aria-label": True}):
            label = elem["aria-label"].strip()
            if 3 < len(label) < 80:
                names.add(label)

    else:
        # Fallback: regex-based extraction if BeautifulSoup not available
        # Match capitalized words that look like company names
        for match in re.finditer(r">([A-Z][A-Za-z0-9\s&\.\-]{2,60})<", html):
            names.add(match.group(1).strip())

    # Filter out obvious non-companies
    filtered = []
    for name in names:
        name = name.strip()
        # Skip if too short or too long
        if len(name) < 3 or len(name) > 80:
            continue
        # Skip if looks like a menu/nav item
        if name.lower() in {"home", "about", "contact", "products", "services", "news", "blog", "login", "register", "search", "menu", "cart", "home page", "about us", "contact us"}:
            continue
        # Skip if mostly numbers
        if sum(c.isdigit() for c in name) > len(name) * 0.5:
            continue
        filtered.append(name)

    return sorted(set(filtered))


# ---------------------------------------------------------------------------
# File reading
# ---------------------------------------------------------------------------

def read_from_excel(path: str, column_name: Optional[str] = None) -> List[str]:
    """Read company names from Excel file."""
    if openpyxl is None:
        print("[ERROR] openpyxl not installed. Run: pip install openpyxl")
        return []

    wb = openpyxl.load_workbook(path)
    ws = wb.active

    # If column name specified, find that column
    if column_name:
        headers = [c.value for c in ws[1]]
        try:
            col_idx = headers.index(column_name)
        except ValueError:
            print(f"[ERROR] Column '{column_name}' not found. Available: {headers}")
            return []
        names = [row[col_idx].value for row in ws.iter_rows(min_row=2, values_only=False)]
    else:
        # Try first column
        names = [row[0].value for row in ws.iter_rows(min_row=2, values_only=False)]

    return [str(n).strip() for n in names if n]


def read_from_csv(path: str, column_name: Optional[str] = None) -> List[str]:
    """Read company names from CSV file."""
    names = []
    with open(path, "r", encoding="utf-8", errors="ignore") as f:
        reader = csv.DictReader(f)
        if column_name:
            for row in reader:
                val = row.get(column_name, "").strip()
                if val:
                    names.append(val)
        else:
            # First column
            for row in reader:
                first = list(row.values())[0].strip() if row else ""
                if first:
                    names.append(first)
    return names


def read_from_text(path: str) -> List[str]:
    """Read company names from a plain text file (one per line)."""
    with open(path, "r", encoding="utf-8", errors="ignore") as f:
        return [line.strip() for line in f if line.strip()]


# ---------------------------------------------------------------------------
# Domain finding via search
# ---------------------------------------------------------------------------

def find_domain_ddg(company_name: str) -> Optional[str]:
    """Try to find the company website domain via DuckDuckGo search."""
    query = f"{company_name} official website"
    try:
        resp = requests.get(
            "https://html.duckduckgo.com/html/",
            params={"q": query},
            headers=HEADERS,
            timeout=20,
        )
        resp.raise_for_status()
        html = resp.text

        # Extract first result URL
        m = re.search(r'<a[^>]+class="result__a"[^>]+href="([^"]+)"', html)
        if not m:
            m = re.search(r'<a[^>]+href="(https?://[^"]+)"', html)
        if not m:
            return None

        url = m.group(1)
        # DuckDuckGo HTML uses redirect URLs
        if "duckduckgo.com/l/" in url or "duckduckgo.com/d.js" in url:
            # Extract actual URL from redirect
            ru = re.search(r'uddg=([^&]+)', url)
            if ru:
                url = urllib.parse.unquote(ru.group(1))

        # Extract domain
        parsed = urllib.parse.urlparse(url)
        domain = parsed.netloc.lower()
        if domain.startswith("www."):
            domain = domain[4:]
        # Skip generic domains
        if domain in {"facebook.com", "linkedin.com", "twitter.com", "x.com", "instagram.com", "youtube.com", "wikipedia.org", "amazon.com", "ebay.com", "alibaba.com", "zoominfo.com", "crunchbase.com", "bbb.org", "yellowpages.com", "yelp.com", "tripadvisor.com", "pinterest.com", "reddit.com"}:
            return None
        return domain
    except Exception as e:
        print(f"    [DDG ERROR] {company_name}: {e}")
        return None


def find_domains(companies: List[str], delay: float = 1.0) -> dict:
    """Find domains for a list of companies."""
    results = {}
    print(f"\n[Domain Search] Finding websites for {len(companies)} companies via DuckDuckGo...")
    for idx, name in enumerate(companies, 1):
        print(f"  [{idx}/{len(companies)}] {name} ...", end=" ", flush=True)
        domain = find_domain_ddg(name)
        if domain:
            results[name] = domain
            print(domain)
        else:
            print("not found")
        time.sleep(delay)
    return results


# ---------------------------------------------------------------------------
# Filtering
# ---------------------------------------------------------------------------

def filter_glove_related(companies: List[str]) -> List[str]:
    """Filter companies that are likely glove-related."""
    matched = [c for c in companies if is_glove_related(c)]
    print(f"\n[Filter] {len(matched)}/{len(companies)} companies match glove keywords")
    if matched:
        print("  Matched:")
        for c in matched[:10]:
            print(f"    - {c}")
        if len(matched) > 10:
            print(f"    ... and {len(matched) - 10} more")
    return matched


# ---------------------------------------------------------------------------
# Export
# ---------------------------------------------------------------------------

def export_csv(companies: List[str], domains: dict, output_path: str):
    """Export results to CSV."""
    with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(["company", "domain", "glove_related"])
        for name in companies:
            domain = domains.get(name, "")
            writer.writerow([name, domain, "yes" if is_glove_related(name) else "no"])
    print(f"\n[OK] Exported {len(companies)} rows to: {output_path}")


def export_domain_list(domains: List[str], output_path: str):
    """Export just domain names (one per line) for easy copy-paste."""
    clean = sorted(set(d for d in domains if d))
    with open(output_path, "w", encoding="utf-8") as f:
        for d in clean:
            f.write(d + "\n")
    print(f"[OK] Exported {len(clean)} domains to: {output_path}")
    print("  Copy these domains and paste into lead-finder's '批量导入域名' field.")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Extract glove-related trade show exhibitors")
    parser.add_argument("--url", type=str, help="Exhibitor list page URL to scrape")
    parser.add_argument("--file", type=str, help="Excel/CSV file path")
    parser.add_argument("--text-file", type=str, help="Plain text file with company names (one per line)")
    parser.add_argument("--col", type=str, help="Column name containing company names (for Excel/CSV)")
    parser.add_argument("--find-domains", action="store_true", help="Search for company website domains via DuckDuckGo")
    parser.add_argument("--delay", type=float, default=1.0, help="Delay between domain searches (seconds)")
    parser.add_argument("--all", action="store_true", help="Include all companies, not just glove-related ones")
    parser.add_argument("--output", type=str, default="trade_show_leads.csv", help="Output CSV path")
    parser.add_argument("--domain-list", type=str, help="Also export a plain domain list file (for bulk import)")
    args = parser.parse_args()

    # Step 1: Collect company names
    companies = []
    if args.url:
        print(f"[Fetch] Downloading {args.url} ...")
        html = fetch_html(args.url)
        companies = extract_company_names_from_html(html)
        print(f"[Extract] Found {len(companies)} potential company names from page")
    elif args.file:
        path = Path(args.file)
        if path.suffix.lower() in {".xlsx", ".xls"}:
            companies = read_from_excel(args.file, args.col)
        elif path.suffix.lower() == ".csv":
            companies = read_from_csv(args.file, args.col)
        else:
            companies = read_from_text(args.file)
        print(f"[Read] Loaded {len(companies)} companies from {args.file}")
    elif args.text_file:
        companies = read_from_text(args.text_file)
        print(f"[Read] Loaded {len(companies)} companies from {args.text_file}")
    else:
        parser.print_help()
        print("\n[ERROR] Please provide --url, --file, or --text-file")
        sys.exit(1)

    if not companies:
        print("[WARNING] No company names found.")
        sys.exit(0)

    # Step 2: Deduplicate and normalize
    companies = sorted(set(normalize_company_name(c) for c in companies if c.strip()))
    print(f"[Deduplicate] {len(companies)} unique companies")

    # Step 3: Filter glove-related
    if args.all:
        filtered = companies
    else:
        filtered = filter_glove_related(companies)

    if not filtered:
        print("[WARNING] No glove-related companies found. Try --all to see all companies, or adjust keywords.")
        sys.exit(0)

    # Step 4: Find domains (optional)
    domains = {}
    if args.find_domains:
        domains = find_domains(filtered, delay=args.delay)
    else:
        print("\n[Tip] Pass --find-domains to automatically search for company websites via DuckDuckGo")

    # Step 5: Export
    export_csv(filtered, domains, args.output)
    if args.domain_list:
        export_domain_list([domains.get(c, "") for c in filtered], args.domain_list)

    # Print summary
    print(f"\n{'='*60}")
    print(f"  SUMMARY")
    print(f"{'='*60}")
    print(f"  Total companies   : {len(companies)}")
    print(f"  Glove-related     : {len(filtered)}")
    print(f"  Domains found     : {sum(1 for d in domains.values() if d)}")
    print(f"  Output CSV        : {args.output}")
    if args.domain_list:
        print(f"  Domain list       : {args.domain_list}")
    print(f"{'='*60}")


if __name__ == "__main__":
    main()
