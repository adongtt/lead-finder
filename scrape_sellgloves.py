#!/usr/bin/env python3
"""Scrape sellgloves.com football gloves products into Excel/CSV."""

import csv
import re
import sys
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from urllib.parse import urljoin

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

BASE_URL = "https://www.sellgloves.com"
LIST_PAGES = [
    "/products/football-gloves.html",
    "/products/football-gloves_2.html",
    "/products/football-gloves_3.html",
    "/products/football-gloves_4.html",
]

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/125.0.0.0 Safari/537.36"
    )
}

session = requests.Session()
session.headers.update(HEADERS)


def log(msg: str):
    # Avoid UnicodeEncodeError on Windows GBK console by encoding fallback
    try:
        print(msg)
    except UnicodeEncodeError:
        print(msg.encode("utf-8", "ignore").decode("ascii", "ignore"))


def fetch(url: str, retries: int = 3) -> str:
    for attempt in range(retries):
        try:
            resp = session.get(url, timeout=30)
            resp.raise_for_status()
            resp.encoding = resp.apparent_encoding
            return resp.text
        except Exception:
            if attempt == retries - 1:
                raise
            time.sleep(1)
    return ""


def parse_list_page(html: str) -> list[dict]:
    soup = BeautifulSoup(html, "html.parser")
    products = []
    for pro in soup.select("div.productlist div.pro"):
        a = pro.find("a", href=True)
        if not a:
            continue
        img = a.find("img")
        grey = a.find("span", class_="grey")
        red = a.find("span", class_="red")

        rel_link = a["href"]
        products.append({
            "title": (a.get("title") or "").strip(),
            "list_image": urljoin(BASE_URL, img["src"]) if img and img.get("src") else "",
            "original_price": grey.get_text(strip=True) if grey else "",
            "sale_price": red.get_text(strip=True) if red else "",
            "product_url": urljoin(BASE_URL, rel_link),
        })
    return products


def parse_detail_page(html: str) -> dict:
    soup = BeautifulSoup(html, "html.parser")

    # ---- Quick Detail ----
    quick_detail: dict[str, str] = {}
    qd_title = soup.find("div", class_="qdtitle")
    if qd_title:
        for sibling in qd_title.find_all_next():
            if sibling.get("class") and "qdtitle" in sibling.get("class", []):
                break
            if sibling.name == "p":
                label_tag = sibling.find("label")
                span_tag = sibling.find("span")
                if label_tag and span_tag:
                    key = label_tag.get_text(strip=True).rstrip(":")
                    val = span_tag.get_text(strip=True)
                    if key:
                        # normalise key immediately
                        norm_key = re.sub(r"\s+", " ", key.strip()).lower()
                        quick_detail[norm_key] = val

    # ---- Product Description ----
    description = ""
    qd_titles = soup.find_all("div", class_="qdtitle")
    if len(qd_titles) >= 2:
        desc_title = qd_titles[1]
        parent = desc_title.find_parent()
        if parent:
            parts = []
            for child in parent.children:
                if getattr(child, "name", None) == "div" and "qdtitle" in child.get("class", []):
                    continue
                txt = child.get_text(separator=" ", strip=True) if hasattr(child, "get_text") else str(child).strip()
                if txt:
                    parts.append(txt)
            description = "\n".join(line.strip() for line in " ".join(parts).splitlines() if line.strip())

    if len(description) < 20:
        spans = soup.find_all("span", style=re.compile(r"white-space:\s*pre-wrap", re.I))
        if spans:
            description = "\n".join(s.get_text(strip=True) for s in spans if s.get_text(strip=True))

    # ---- Images ----
    # product gallery images (bigphoto + thumbnail links)
    image_set: set[str] = set()
    bigphoto = soup.select_one("div.bigphoto img")
    if bigphoto and bigphoto.get("src"):
        image_set.add(urljoin(BASE_URL, bigphoto["src"]))

    # Thumbnail links usually in the same productdetail area
    for a_tag in soup.select("div.productdetail a"):
        img = a_tag.find("img")
        if img and img.get("src"):
            src = urljoin(BASE_URL, img["src"])
            # filter out generic UI images
            if "/uploadfile/" in src or "/products/" in src.lower():
                image_set.add(src)

    detail_images = ", ".join(sorted(image_set)) if image_set else ""

    return {
        "quick_detail": quick_detail,
        "description": description,
        "detail_images": detail_images,
    }


def scrape_one_product(product: dict) -> dict:
    try:
        html = fetch(product["product_url"])
        detail = parse_detail_page(html)
        product.update(detail)
    except Exception as exc:
        log(f"  [ERR] {product['product_url']} -> {exc}")
        product["quick_detail"] = {}
        product["description"] = ""
        product["detail_images"] = ""
    return product


def write_excel(results: list[dict], path: str):
    # Normalise all quick-detail keys across rows
    all_qd_keys: set[str] = set()
    for r in results:
        all_qd_keys.update(r.get("quick_detail", {}).keys())
    qd_keys = sorted(all_qd_keys)

    wb = Workbook()
    ws = wb.active
    ws.title = "Football Gloves"

    headers = [
        "Title",
        "Original Price",
        "Sale Price",
        "Description",
        *[k.title() for k in qd_keys],
        "List Image",
        "Detail Images",
        "Product URL",
    ]

    # Header row style
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Data rows
    for row_idx, r in enumerate(results, 2):
        row_data = [
            r.get("title", ""),
            r.get("original_price", ""),
            r.get("sale_price", ""),
            r.get("description", ""),
            *[r.get("quick_detail", {}).get(k, "") for k in qd_keys],
            r.get("list_image", ""),
            r.get("detail_images", ""),
            r.get("product_url", ""),
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # Auto-width (rough) + freeze header
    for col_idx, h in enumerate(headers, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max(len(str(h)) + 2, 12), 50)
    ws.freeze_panes = "A2"

    wb.save(path)


def write_csv(results: list[dict], path: str):
    all_qd_keys: set[str] = set()
    for r in results:
        all_qd_keys.update(r.get("quick_detail", {}).keys())
    qd_keys = sorted(all_qd_keys)

    fieldnames = [
        "title",
        "original_price",
        "sale_price",
        "description",
        *[k for k in qd_keys],
        "list_image",
        "detail_images",
        "product_url",
    ]
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for r in results:
            row = {
                "title": r.get("title", ""),
                "original_price": r.get("original_price", ""),
                "sale_price": r.get("sale_price", ""),
                "description": r.get("description", ""),
                "list_image": r.get("list_image", ""),
                "detail_images": r.get("detail_images", ""),
                "product_url": r.get("product_url", ""),
            }
            for k in qd_keys:
                row[k] = r.get("quick_detail", {}).get(k, "")
            writer.writerow(row)


def main():
    log("Step 1/3: Fetching list pages...")
    all_products: list[dict] = []
    for path in LIST_PAGES:
        url = urljoin(BASE_URL, path)
        log(f"  -> {url}")
        html = fetch(url)
        items = parse_list_page(html)
        log(f"     Found {len(items)} products")
        all_products.extend(items)
    log(f"Total products to scrape: {len(all_products)}")

    log("\nStep 2/3: Fetching detail pages (parallel)...")
    results: list[dict] = []
    done = 0
    with ThreadPoolExecutor(max_workers=10) as executor:
        futures = {executor.submit(scrape_one_product, p): p for p in all_products}
        for future in as_completed(futures):
            results.append(future.result())
            done += 1
            if done % 10 == 0 or done == len(all_products):
                log(f"  ... {done}/{len(all_products)} done")

    log("\nStep 3/3: Writing files...")
    write_excel(results, "sellgloves_football_gloves.xlsx")
    write_csv(results, "sellgloves_football_gloves.csv")
    log(f"Saved: sellgloves_football_gloves.xlsx  ({len(results)} rows)")
    log(f"Saved: sellgloves_football_gloves.csv   ({len(results)} rows)")


if __name__ == "__main__":
    main()
